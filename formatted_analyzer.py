#!/usr/bin/env python3
"""
Enhanced Weekly Attendance Analyzer with Professional Excel Formatting
Based on analysis of manually formatted Excel file to preserve professional appearance
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
import os
import warnings

warnings.filterwarnings('ignore')

class FormattedAttendanceAnalyzer:
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.data = {}
        self.attendance_data = pd.DataFrame()
        self.week_start = "31-Aug-2025"  # Week start date
        
        # Professional formatting styles based on analysis
        self.header_font = Font(name='Calibri', size=12, bold=True)
        self.header_alignment = Alignment(horizontal='center', vertical='center')
        self.data_alignment = Alignment(horizontal='center', vertical='center')
        self.header_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Custom column widths from analysis
        self.column_widths = {
            'Summary': {
                'A': 14.05, 'B': 16.84, 'C': 17.73, 
                'D': 18.52, 'E': 18.16, 'F': 23.58
            },
            'All Students': {
                'A': 14.05, 'B': 18.63, 'C': 32.42, 'D': 13.68,
                'E': 17.16, 'F': 16.37, 'G': 16.47, 'H': 15.52,
                'I': 15.10, 'J': 14.31, 'K': 15.16, 'L': 14.37,
                'M': 15.37, 'N': 16.16, 'O': 15.37, 'P': 16.21, 'Q': 15.42
            },
            'Full Week Students': {
                'A': 10.21, 'B': 18.63, 'C': 28.26, 'D': 13.68
            },
            'Never Attended': {
                'A': 15.26, 'B': 18.63, 'C': 26.89, 'D': 13.68
            }
        }
        
        self.custom_row_height = 15.6
        
    def read_excel_data(self):
        """Read data from all sheets in the Excel file"""
        print("Reading Excel data...")
        try:
            excel_data = pd.ExcelFile(self.excel_file)
            print(f"Found {len(excel_data.sheet_names)} sheets")
            
            for sheet_name in excel_data.sheet_names:
                print(f"Processing sheet: {sheet_name}")
                
                # Read data starting from row 4 (index 3) to skip headers
                df = pd.read_excel(
                    self.excel_file, 
                    sheet_name=sheet_name, 
                    header=None,
                    skiprows=3
                )
                
                self.data[sheet_name] = df
                
        except Exception as e:
            print(f"Error reading Excel file: {str(e)}")
            return False
        
        return True
    
    def process_attendance_data(self):
        """Process attendance data from all sheets"""
        print("\nProcessing attendance data...")
        
        all_students = []
        
        for sheet_name, df in self.data.items():
            print(f"Processing sheet: {sheet_name}")
            
            # Skip if dataframe is empty or too small
            if df.empty or df.shape[0] < 2:
                print(f"Skipping {sheet_name} - insufficient data")
                continue
            
            # Find student data (names in column B, index 1)
            for idx, row in df.iterrows():
                student_name = row.iloc[1] if len(row) > 1 else None
                
                if pd.isna(student_name) or str(student_name).strip() == '':
                    continue
                
                # Skip if it looks like a header or metadata
                if any(keyword in str(student_name).lower() for keyword in 
                       ['اسم', 'name', 'total', 'المجموع', 'group']):
                    continue
                
                student_number = row.iloc[0] if len(row) > 0 and not pd.isna(row.iloc[0]) else 'N/A'
                student_id = row.iloc[2] if len(row) > 2 and not pd.isna(row.iloc[2]) else 'N/A'
                
                # Extract attendance data (columns D onwards, index 3+)
                attendance_sessions = []
                for col_idx in range(3, min(len(row), 23)):  # Up to 20 sessions
                    session_value = row.iloc[col_idx]
                    if pd.isna(session_value):
                        attendance_sessions.append(0)
                    else:
                        attendance_sessions.append(1 if float(session_value) >= 1.0 else 0)
                
                if len(attendance_sessions) < 20:
                    attendance_sessions.extend([0] * (20 - len(attendance_sessions)))
                
                # Calculate daily attendance (3+ out of 4 sessions = present)
                daily_attendance = []
                for day in range(5):  # 5 days
                    start_session = day * 4
                    end_session = start_session + 4
                    day_sessions = attendance_sessions[start_session:end_session]
                    daily_present = 1 if sum(day_sessions) >= 3 else 0
                    daily_attendance.append(daily_present)
                
                # Calculate statistics
                days_attended = sum(daily_attendance)
                total_sessions = sum(attendance_sessions)
                attendance_percentage = (days_attended / 5) * 100
                
                student_record = {
                    'Group': sheet_name,
                    'Student Number': student_number,
                    'Student Name': str(student_name).strip(),
                    'Student ID': student_id,
                    'Days Attended': days_attended,
                    'Attendance %': attendance_percentage,
                    'Total Sessions': total_sessions,
                    'Daily Attendance': daily_attendance
                }
                
                # Add daily details
                days = ['Sun (31-Aug)', 'Mon (1-Sep)', 'Tue (2-Sep)', 'Wed (3-Sep)', 'Thu (4-Sep)']
                for i, day in enumerate(days):
                    student_record[day] = 'Present' if daily_attendance[i] == 1 else 'Absent'
                
                all_students.append(student_record)
        
        self.attendance_data = pd.DataFrame(all_students)
        print(f"Processed {len(all_students)} students")
        return True
    
    def apply_professional_formatting(self, worksheet, sheet_name):
        """Apply professional formatting matching the manual formatting"""
        
        # Set column widths
        if sheet_name in self.column_widths:
            for col_letter, width in self.column_widths[sheet_name].items():
                worksheet.column_dimensions[col_letter].width = width
        
        # Format header row
        for cell in worksheet[1]:
            if cell.value:
                cell.font = self.header_font
                cell.alignment = self.header_alignment
                cell.border = self.header_border
        
        # Set row heights
        for row_num in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[row_num].height = self.custom_row_height
        
        # Apply data alignment for data rows
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                if cell.value is not None:
                    cell.alignment = self.data_alignment
    
    def create_formatted_excel_report(self):
        """Create Excel report with professional formatting"""
        
        # Generate week date range
        from datetime import datetime
        start_date = datetime.strptime("31-08-2025", "%d-%m-%Y")
        end_date = start_date + timedelta(days=4)
        date_range = f"{start_date.strftime('%d%b')}-{end_date.strftime('%d%b')}"
        
        output_file = f"weekly_attendance_results_{date_range}.xlsx"
        
        print(f"\nCreating formatted Excel report: {output_file}")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            
            # Summary Sheet
            summary_data = []
            for group in self.attendance_data['Group'].unique():
                group_data = self.attendance_data[self.attendance_data['Group'] == group]
                total_students = len(group_data)
                full_week = len(group_data[group_data['Days Attended'] == 5])
                partial = len(group_data[(group_data['Days Attended'] > 0) & (group_data['Days Attended'] < 5)])
                never_attended = len(group_data[group_data['Days Attended'] == 0])
                avg_attendance = group_data['Attendance %'].mean()
                
                summary_data.append({
                    'Group': group,
                    'Total Students': total_students,
                    'Full Week (5/5)': full_week,
                    'Partial (1-4)': partial,
                    'Never Attended': never_attended,
                    'Average Attendance %': round(avg_attendance, 1)
                })
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # All Students Sheet
            columns_order = ['Group', 'Student Number', 'Student Name', 'Student ID', 
                           'Days Attended', 'Attendance %', 'Total Sessions',
                           'Sun (31-Aug)', 'Mon (1-Sep)', 'Tue (2-Sep)', 'Wed (3-Sep)', 'Thu (4-Sep)']
            
            all_students_df = self.attendance_data[columns_order].copy()
            all_students_df.to_excel(writer, sheet_name='All Students', index=False)
            
            # Full Week Students
            full_week_students = self.attendance_data[self.attendance_data['Days Attended'] == 5]
            full_week_df = full_week_students[['Group', 'Student Number', 'Student Name', 'Student ID']].copy()
            full_week_df.to_excel(writer, sheet_name='Full Week Students', index=False)
            
            # Never Attended Students
            never_attended = self.attendance_data[self.attendance_data['Days Attended'] == 0]
            never_df = never_attended[['Group', 'Student Number', 'Student Name', 'Student ID']].copy()
            never_df.to_excel(writer, sheet_name='Never Attended', index=False)
        
        # Apply professional formatting to each sheet
        wb = openpyxl.load_workbook(output_file)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self.apply_professional_formatting(ws, sheet_name)
            
            # Apply conditional formatting for percentage columns
            if sheet_name == 'Summary':
                # Color scale for Average Attendance %
                color_scale_rule = ColorScaleRule(
                    start_type='min', start_color='FFE6E6',
                    end_type='max', end_color='E6F3E6'
                )
                ws.conditional_formatting.add('F2:F100', color_scale_rule)
                
            elif sheet_name == 'All Students':
                # Color scale for Days Attended
                color_scale_rule1 = ColorScaleRule(
                    start_type='min', start_color='FFE6E6',
                    end_type='max', end_color='E6F3E6'
                )
                ws.conditional_formatting.add('E2:E1000', color_scale_rule1)
                
                # Color scale for Attendance %
                color_scale_rule2 = ColorScaleRule(
                    start_type='min', start_color='FFE6E6',
                    end_type='max', end_color='E6F3E6'
                )
                ws.conditional_formatting.add('F2:F1000', color_scale_rule2)
        
        wb.save(output_file)
        wb.close()
        
        print(f"✅ Formatted Excel report saved: {output_file}")
        return output_file

    def create_power_bi_style_visualizations(self):
        """Create Power BI-style visualizations with professional formatting"""
        print("\nCreating Power BI-style visualizations...")
        
        # Set professional style
        plt.style.use('default')
        sns.set_palette("husl")
        
        # 1. Large Attendance Overview Chart
        fig, ax = plt.subplots(figsize=(20, 12))
        
        # Calculate overall statistics
        total_students = len(self.attendance_data)
        full_week_count = len(self.attendance_data[self.attendance_data['Days Attended'] == 5])
        partial_count = len(self.attendance_data[(self.attendance_data['Days Attended'] > 0) & 
                                                (self.attendance_data['Days Attended'] < 5)])
        never_attended_count = len(self.attendance_data[self.attendance_data['Days Attended'] == 0])
        
        categories = ['Full Week\n(5/5 days)', 'Partial\n(1-4 days)', 'Never\nAttended']
        values = [full_week_count, partial_count, never_attended_count]
        colors = ['#2E8B57', '#FFD700', '#DC143C']
        
        bars = ax.bar(categories, values, color=colors, alpha=0.8, edgecolor='white', linewidth=2)
        
        # Add value labels on bars
        for i, bar in enumerate(bars):
            height = bar.get_height()
            percentage = (values[i] / total_students) * 100
            ax.text(bar.get_x() + bar.get_width()/2., height + 2,
                   f'{int(height)}\n({percentage:.1f}%)',
                   ha='center', va='bottom', fontsize=16, fontweight='bold')
        
        ax.set_title('Weekly Attendance Overview (31 Aug - 4 Sep 2025)', fontsize=24, fontweight='bold', pad=30)
        ax.set_ylabel('Number of Students', fontsize=18, fontweight='bold')
        ax.set_ylim(0, max(values) * 1.2)
        ax.tick_params(axis='both', which='major', labelsize=14)
        ax.grid(True, alpha=0.3, axis='y')
        
        plt.tight_layout()
        plt.savefig('attendance_overview_large.png', dpi=300, bbox_inches='tight', facecolor='white')
        plt.close()
        
        # 2. Large Group Performance Chart
        fig, ax = plt.subplots(figsize=(18, 10))
        
        group_stats = []
        for group in sorted(self.attendance_data['Group'].unique()):
            group_data = self.attendance_data[self.attendance_data['Group'] == group]
            avg_attendance = group_data['Attendance %'].mean()
            group_stats.append({'Group': group, 'Average Attendance': avg_attendance})
        
        group_df = pd.DataFrame(group_stats)
        
        bars = ax.bar(range(len(group_df)), group_df['Average Attendance'], 
                     color='steelblue', alpha=0.7, edgecolor='navy', linewidth=1)
        
        # Add percentage labels
        for i, bar in enumerate(bars):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height + 1,
                   f'{height:.1f}%', ha='center', va='bottom', fontsize=12, fontweight='bold')
        
        ax.set_title('Average Attendance by Group (31 Aug - 4 Sep 2025)', fontsize=20, fontweight='bold', pad=20)
        ax.set_xlabel('Groups', fontsize=16, fontweight='bold')
        ax.set_ylabel('Average Attendance %', fontsize=16, fontweight='bold')
        ax.set_xticks(range(len(group_df)))
        ax.set_xticklabels(group_df['Group'], rotation=45, ha='right')
        ax.set_ylim(0, 100)
        ax.tick_params(axis='both', which='major', labelsize=12)
        ax.grid(True, alpha=0.3, axis='y')
        
        plt.tight_layout()
        plt.savefig('group_performance_large.png', dpi=300, bbox_inches='tight', facecolor='white')
        plt.close()
        
        # 3. Large Daily Attendance Trend
        fig, ax = plt.subplots(figsize=(16, 10))
        
        days = ['Sun (31-Aug)', 'Mon (1-Sep)', 'Tue (2-Sep)', 'Wed (3-Sep)', 'Thu (4-Sep)']
        daily_counts = []
        
        for day in days:
            present_count = len(self.attendance_data[self.attendance_data[day] == 'Present'])
            daily_counts.append(present_count)
        
        ax.plot(days, daily_counts, marker='o', linewidth=4, markersize=10, 
                color='darkgreen', markerfacecolor='lightgreen', markeredgewidth=2)
        ax.fill_between(days, daily_counts, alpha=0.3, color='lightgreen')
        
        # Add value labels
        for i, count in enumerate(daily_counts):
            percentage = (count / total_students) * 100
            ax.text(i, count + 5, f'{count}\n({percentage:.1f}%)', 
                   ha='center', va='bottom', fontsize=12, fontweight='bold')
        
        ax.set_title('Daily Attendance Trend (31 Aug - 4 Sep 2025)', fontsize=20, fontweight='bold', pad=20)
        ax.set_xlabel('Days', fontsize=16, fontweight='bold')
        ax.set_ylabel('Students Present', fontsize=16, fontweight='bold')
        ax.tick_params(axis='both', which='major', labelsize=12)
        ax.grid(True, alpha=0.3)
        ax.set_ylim(0, max(daily_counts) * 1.1)
        
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        plt.savefig('daily_trend_large.png', dpi=300, bbox_inches='tight', facecolor='white')
        plt.close()
        
        # 4. Large Attendance Distribution Pie Chart
        fig, ax = plt.subplots(figsize=(14, 14))
        
        pie_labels = ['Full Week (5/5)', 'Partial (1-4)', 'Never Attended']
        pie_values = [full_week_count, partial_count, never_attended_count]
        pie_colors = ['#2E8B57', '#FFD700', '#DC143C']
        
        wedges, texts, autotexts = ax.pie(pie_values, labels=pie_labels, colors=pie_colors,
                                         autopct=lambda pct: f'{pct:.1f}%\n({int(pct/100*total_students)})',
                                         startangle=90, textprops={'fontsize': 14, 'fontweight': 'bold'})
        
        ax.set_title('Attendance Distribution (31 Aug - 4 Sep 2025)', 
                    fontsize=20, fontweight='bold', pad=30)
        
        plt.tight_layout()
        plt.savefig('attendance_distribution_large.png', dpi=300, bbox_inches='tight', facecolor='white')
        plt.close()
        
        print("✅ Power BI-style visualizations created:")
        print("  - attendance_overview_large.png")
        print("  - group_performance_large.png")
        print("  - daily_trend_large.png")
        print("  - attendance_distribution_large.png")

    def create_professional_html_dashboard(self):
        """Create professional HTML dashboard with embedded charts"""
        
        total_students = len(self.attendance_data)
        full_week_count = len(self.attendance_data[self.attendance_data['Days Attended'] == 5])
        partial_count = len(self.attendance_data[(self.attendance_data['Days Attended'] > 0) & 
                                                (self.attendance_data['Days Attended'] < 5)])
        never_attended_count = len(self.attendance_data[self.attendance_data['Days Attended'] == 0])
        
        html_content = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Weekly Attendance Dashboard - 31 Aug to 4 Sep 2025</title>
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 0; padding: 20px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); }}
        .container {{ max-width: 1400px; margin: 0 auto; background: white; border-radius: 15px; box-shadow: 0 10px 30px rgba(0,0,0,0.2); padding: 30px; }}
        h1 {{ color: #2c3e50; text-align: center; font-size: 2.5em; margin-bottom: 10px; }}
        .date-range {{ text-align: center; color: #7f8c8d; font-size: 1.2em; margin-bottom: 40px; }}
        .stats-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 20px; margin-bottom: 40px; }}
        .stat-card {{ background: linear-gradient(145deg, #f8f9fa, #e9ecef); padding: 25px; border-radius: 12px; text-align: center; box-shadow: 0 5px 15px rgba(0,0,0,0.1); }}
        .stat-number {{ font-size: 3em; font-weight: bold; color: #2c3e50; }}
        .stat-label {{ font-size: 1.1em; color: #6c757d; margin-top: 10px; }}
        .chart-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 30px; margin: 30px 0; }}
        .chart-container {{ text-align: center; background: #f8f9fa; padding: 20px; border-radius: 12px; }}
        .chart-container img {{ max-width: 100%; height: auto; border-radius: 8px; }}
        .full-width {{ grid-column: 1 / -1; }}
        .summary-section {{ margin: 40px 0; }}
        .summary-table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        .summary-table th, .summary-table td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }}
        .summary-table th {{ background-color: #3498db; color: white; }}
        .summary-table tr:hover {{ background-color: #f5f5f5; }}
        .footer {{ text-align: center; margin-top: 40px; padding: 20px; background: #ecf0f1; border-radius: 8px; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>📊 Weekly Attendance Dashboard</h1>
        <p class="date-range">Week: 31 August - 4 September 2025</p>
        
        <div class="stats-grid">
            <div class="stat-card">
                <div class="stat-number">{total_students}</div>
                <div class="stat-label">Total Students</div>
            </div>
            <div class="stat-card">
                <div class="stat-number" style="color: #27ae60;">{full_week_count}</div>
                <div class="stat-label">Full Week (5/5)</div>
            </div>
            <div class="stat-card">
                <div class="stat-number" style="color: #f39c12;">{partial_count}</div>
                <div class="stat-label">Partial (1-4)</div>
            </div>
            <div class="stat-card">
                <div class="stat-number" style="color: #e74c3c;">{never_attended_count}</div>
                <div class="stat-label">Never Attended</div>
            </div>
        </div>

        <div class="chart-grid">
            <div class="chart-container">
                <h3>📈 Attendance Overview</h3>
                <img src="attendance_overview_large.png" alt="Attendance Overview">
            </div>
            <div class="chart-container">
                <h3>📊 Group Performance</h3>
                <img src="group_performance_large.png" alt="Group Performance">
            </div>
        </div>

        <div class="chart-grid">
            <div class="chart-container">
                <h3>📅 Daily Attendance Trend</h3>
                <img src="daily_trend_large.png" alt="Daily Trend">
            </div>
            <div class="chart-container">
                <h3>🥧 Attendance Distribution</h3>
                <img src="attendance_distribution_large.png" alt="Attendance Distribution">
            </div>
        </div>

        <div class="summary-section">
            <h3>📋 Group Summary</h3>
            <table class="summary-table">
                <thead>
                    <tr>
                        <th>Group</th>
                        <th>Total Students</th>
                        <th>Full Week (5/5)</th>
                        <th>Partial (1-4)</th>
                        <th>Never Attended</th>
                        <th>Average Attendance %</th>
                    </tr>
                </thead>
                <tbody>
"""

        # Add group summary rows
        for group in sorted(self.attendance_data['Group'].unique()):
            group_data = self.attendance_data[self.attendance_data['Group'] == group]
            group_total = len(group_data)
            group_full = len(group_data[group_data['Days Attended'] == 5])
            group_partial = len(group_data[(group_data['Days Attended'] > 0) & (group_data['Days Attended'] < 5)])
            group_never = len(group_data[group_data['Days Attended'] == 0])
            group_avg = group_data['Attendance %'].mean()
            
            html_content += f"""
                    <tr>
                        <td><strong>{group}</strong></td>
                        <td>{group_total}</td>
                        <td>{group_full}</td>
                        <td>{group_partial}</td>
                        <td>{group_never}</td>
                        <td>{group_avg:.1f}%</td>
                    </tr>"""

        html_content += f"""
                </tbody>
            </table>
        </div>

        <div class="footer">
            <p>📅 Report generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            <p>🎯 Professional attendance analysis with preserved formatting</p>
        </div>
    </div>
</body>
</html>"""

        with open('professional_attendance_dashboard.html', 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print("✅ Professional HTML dashboard created: professional_attendance_dashboard.html")

    def run_complete_analysis(self):
        """Run the complete formatted analysis"""
        print("🚀 Starting Professional Attendance Analysis with Preserved Formatting")
        print("=" * 80)
        
        if not self.read_excel_data():
            return False
        
        if not self.process_attendance_data():
            return False
        
        if self.attendance_data.empty:
            print("❌ No valid attendance data found")
            return False
        
        print(f"\n📊 Analysis Summary:")
        print(f"   • Total Students: {len(self.attendance_data)}")
        print(f"   • Groups: {len(self.attendance_data['Group'].unique())}")
        print(f"   • Week: 31 August - 4 September 2025")
        
        # Generate all outputs with professional formatting
        excel_file = self.create_formatted_excel_report()
        self.create_power_bi_style_visualizations()
        self.create_professional_html_dashboard()
        
        print("\n" + "=" * 80)
        print("✅ PROFESSIONAL ANALYSIS COMPLETE!")
        print("\nGenerated files with preserved formatting:")
        print(f"📊 Excel Report: {excel_file}")
        print("📈 Charts: attendance_overview_large.png, group_performance_large.png")
        print("📅        daily_trend_large.png, attendance_distribution_large.png")
        print("🌐 Dashboard: professional_attendance_dashboard.html")
        
        return True

def main():
    """Main function to run the formatted analyzer"""
    excel_file = r'f:\work\Black Gold\attendance statistics\Attendance sheets\كشوفات الغياب الاسبوعي 31-8-2025(drive).xlsx'
    
    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return
    
    analyzer = FormattedAttendanceAnalyzer(excel_file)
    analyzer.run_complete_analysis()

if __name__ == "__main__":
    main()
